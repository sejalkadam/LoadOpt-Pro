"""Export load plan to Excel and PDF."""

import io
from collections import Counter, defaultdict
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


def _thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _container_contents_summary(container):
    """Return list of (box_name, count, total_weight) sorted by count desc."""
    counts = defaultdict(lambda: {"count": 0, "weight": 0.0})
    for b in container["placed_boxes"]:
        counts[b["name"]]["count"] += 1
        counts[b["name"]]["weight"] += b["weight"]
    return sorted(counts.items(), key=lambda x: -x[1]["count"])


def export_excel(data: dict) -> io.BytesIO:
    result = data.get("result", data)
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # ── Styles ────────────────────────────────────────────────
    header_font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1B3A5C")
    sub_fill     = PatternFill("solid", fgColor="E8EEF4")
    accent_fill  = PatternFill("solid", fgColor="2E6DA4")
    section_fill = PatternFill("solid", fgColor="D0E4F7")
    normal_font  = Font(name="Calibri", size=10)
    bold_font    = Font(name="Calibri", bold=True, size=10)
    small_font   = Font(name="Calibri", size=9)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    def hdr_cell(ws, row, col, val, fill=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = header_font
        c.fill = fill or header_fill
        c.alignment = center
        c.border = _thin_border()
        return c

    def data_cell(ws, row, col, val, bold=False, align=center, font=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = font or (bold_font if bold else normal_font)
        c.alignment = align
        c.border = _thin_border()
        return c

    # ── Summary Sheet ──────────────────────────────────────────
    ws_summary.merge_cells("A1:I1")
    t = ws_summary["A1"]
    t.value = "LOAD OPTIMISATION REPORT"
    t.font = Font(name="Calibri", bold=True, size=16, color="1B3A5C")
    t.alignment = center
    ws_summary.row_dimensions[1].height = 36

    ws_summary.merge_cells("A2:I2")
    d = ws_summary["A2"]
    d.value = f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}"
    d.font = Font(name="Calibri", size=10, color="666666")
    d.alignment = center

    # KPIs
    kpi_row = 4
    kpis = [
        ("Total Boxes", result["total_boxes"]),
        ("Total Weight (kg)", result["total_weight"]),
        ("Total Volume (mm³)", f"{result['total_volume']:,.0f}"),
        ("Containers Used", result["num_containers"]),
        ("Unplaced Boxes", len(result.get("unplaced", []))),
    ]
    for i, (label, val) in enumerate(kpis):
        hdr_cell(ws_summary, kpi_row, i + 1, label)
        data_cell(ws_summary, kpi_row + 1, i + 1, val, bold=True)
    ws_summary.row_dimensions[kpi_row].height = 22
    ws_summary.row_dimensions[kpi_row + 1].height = 22

    # ── Container Breakdown Table ──────────────────────────────
    row = kpi_row + 3
    ws_summary.merge_cells(f"A{row}:I{row}")
    ws_summary[f"A{row}"].value = "CONTAINER BREAKDOWN"
    ws_summary[f"A{row}"].font = Font(name="Calibri", bold=True, size=12, color="1B3A5C")

    row += 1
    cols_c = ["#", "Container", "Dimensions (L×W×H mm)", "Max Wt (kg)", "Loaded Wt (kg)",
              "Wt Util %", "Vol Util %", "Boxes", "Status"]
    for ci, h in enumerate(cols_c, 1):
        hdr_cell(ws_summary, row, ci, h)

    for idx, c in enumerate(result["containers"], 1):
        row += 1
        cd = c["container_dims"]
        data_cell(ws_summary, row, 1, idx)
        data_cell(ws_summary, row, 2, c["container_name"], align=left)
        data_cell(ws_summary, row, 3, f"{cd['l']}×{cd['w']}×{cd['h']}")
        data_cell(ws_summary, row, 4, c["max_wt"])
        data_cell(ws_summary, row, 5, c["total_weight"])

        util_wt = c["utilization_wt"]
        wc = ws_summary.cell(row=row, column=6, value=util_wt)
        wc.font = normal_font; wc.alignment = center; wc.border = _thin_border()
        if util_wt > 90:   wc.fill = PatternFill("solid", fgColor="FFB3B3")
        elif util_wt > 70: wc.fill = PatternFill("solid", fgColor="FFE5A0")
        else:              wc.fill = PatternFill("solid", fgColor="B3FFCC")

        data_cell(ws_summary, row, 7, c["utilization_vol"])
        data_cell(ws_summary, row, 8, c["box_count"])

        # Status: how full
        wt_u = c["utilization_wt"]
        status = "Heavy Load" if wt_u > 90 else ("Moderate Load" if wt_u > 60 else "Light Load")
        sc = ws_summary.cell(row=row, column=9, value=status)
        sc.font = normal_font; sc.alignment = center; sc.border = _thin_border()

    # ── Container Contents Summary ─────────────────────────────
    row += 2
    ws_summary.merge_cells(f"A{row}:I{row}")
    ws_summary[f"A{row}"].value = "WHAT'S IN EACH CONTAINER"
    ws_summary[f"A{row}"].font = Font(name="Calibri", bold=True, size=12, color="1B3A5C")

    for cidx, c in enumerate(result["containers"], 1):
        row += 1
        # Container header row
        ws_summary.merge_cells(f"A{row}:I{row}")
        cell = ws_summary[f"A{row}"]
        cd = c["container_dims"]
        cell.value = (f"  Container {cidx}: {c['container_name']}  |  "
                      f"{c['box_count']} boxes  |  "
                      f"Weight: {c['total_weight']:,.1f} kg / {c['max_wt']:,} kg  |  "
                      f"Wt Util: {c['utilization_wt']:.1f}%  |  Vol Util: {c['utilization_vol']:.1f}%")
        cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E6DA4")
        cell.alignment = left
        cell.border = _thin_border()
        ws_summary.row_dimensions[row].height = 20

        # Sub-header for box breakdown
        row += 1
        sub_hdrs = ["", "Box / Item Name", "Qty in this Container", "Weight Each (kg)", "Total Weight (kg)", "% of Container Wt", "", "", ""]
        for ci, h in enumerate(sub_hdrs, 1):
            c2 = ws_summary.cell(row=row, column=ci, value=h)
            c2.font = Font(name="Calibri", bold=True, size=9, color="1B3A5C")
            c2.fill = PatternFill("solid", fgColor="D0E4F7")
            c2.alignment = center
            c2.border = _thin_border()

        contents = _container_contents_summary(c)
        for box_name, info in contents:
            row += 1
            qty   = info["count"]
            tw    = info["weight"]
            wt_ea = round(tw / qty, 3) if qty else 0
            pct   = round(tw / c["total_weight"] * 100, 1) if c["total_weight"] else 0

            ws_summary.cell(row=row, column=1, value="").border = _thin_border()
            data_cell(ws_summary, row, 2, box_name, align=left)
            data_cell(ws_summary, row, 3, qty)
            data_cell(ws_summary, row, 4, wt_ea)
            data_cell(ws_summary, row, 5, round(tw, 2))

            pc = ws_summary.cell(row=row, column=6, value=pct)
            pc.font = normal_font; pc.alignment = center; pc.border = _thin_border()
            if pct > 50: pc.fill = PatternFill("solid", fgColor="FFE5A0")

            for ci in [7, 8, 9]:
                ws_summary.cell(row=row, column=ci, value="").border = _thin_border()

        row += 1  # blank spacer row

    # set column widths
    for ci, w in enumerate([4, 32, 18, 15, 18, 14, 14, 8, 14], 1):
        ws_summary.column_dimensions[get_column_letter(ci)].width = w

    # ── Per-container detail sheets ────────────────────────────
    for cidx, c in enumerate(result["containers"], 1):
        ws = wb.create_sheet(title=f"Container {cidx}")

        # Title
        ws.merge_cells("A1:J1")
        ws["A1"].value = f"Container {cidx}: {c['container_name']}"
        ws["A1"].font = Font(name="Calibri", bold=True, size=13, color="1B3A5C")
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 28

        # Stats bar
        cd = c["container_dims"]
        ws.merge_cells("A2:J2")
        ws["A2"].value = (f"Dims: {cd['l']}×{cd['w']}×{cd['h']} mm  |  "
                         f"Boxes: {c['box_count']}  |  "
                         f"Loaded Wt: {c['total_weight']:,.1f} kg / {c['max_wt']:,} kg  |  "
                         f"Wt Utilisation: {c['utilization_wt']:.1f}%  |  "
                         f"Vol Utilisation: {c['utilization_vol']:.1f}%")
        ws["A2"].font = Font(name="Calibri", size=9, color="FFFFFF")
        ws["A2"].fill = PatternFill("solid", fgColor="2E6DA4")
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 18

        # Contents summary mini-table
        ws.merge_cells("A3:J3")
        ws["A3"].value = "CONTENTS SUMMARY"
        ws["A3"].font = Font(name="Calibri", bold=True, size=10, color="1B3A5C")
        ws["A3"].fill = PatternFill("solid", fgColor="D0E4F7")
        ws["A3"].alignment = center
        ws.row_dimensions[3].height = 16

        sum_hdrs = ["Box / Item Name", "Qty", "Wt Each (kg)", "Total Wt (kg)", "% of Load"]
        for ci2, h in enumerate(sum_hdrs, 1):
            c2 = ws.cell(row=4, column=ci2, value=h)
            c2.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
            c2.fill = PatternFill("solid", fgColor="1B3A5C")
            c2.alignment = center
            c2.border = _thin_border()

        contents = _container_contents_summary(c)
        sum_row = 5
        for box_name, info in contents:
            qty   = info["count"]
            tw    = info["weight"]
            wt_ea = round(tw / qty, 3) if qty else 0
            pct   = round(tw / c["total_weight"] * 100, 1) if c["total_weight"] else 0
            data_cell(ws, sum_row, 1, box_name, align=left)
            data_cell(ws, sum_row, 2, qty)
            data_cell(ws, sum_row, 3, wt_ea)
            data_cell(ws, sum_row, 4, round(tw, 2))
            data_cell(ws, sum_row, 5, pct)
            for ci2 in range(6, 11):
                ws.cell(row=sum_row, column=ci2, value="").border = _thin_border()
            sum_row += 1

        # Blank separator
        blank_row = sum_row
        ws.merge_cells(f"A{blank_row}:J{blank_row}")
        ws.row_dimensions[blank_row].height = 8

        # Box detail header
        detail_hdr_row = blank_row + 1
        ws.merge_cells(f"A{detail_hdr_row}:J{detail_hdr_row}")
        ws[f"A{detail_hdr_row}"].value = "FULL PLACEMENT DETAILS"
        ws[f"A{detail_hdr_row}"].font = Font(name="Calibri", bold=True, size=10, color="1B3A5C")
        ws[f"A{detail_hdr_row}"].fill = PatternFill("solid", fgColor="D0E4F7")
        ws[f"A{detail_hdr_row}"].alignment = center
        ws.row_dimensions[detail_hdr_row].height = 16

        hdrs = ["#", "Box Name", "Color", "Rotation", "X (mm)", "Y (mm)", "Z (mm)",
                "Dims (dx×dy×dz mm)", "Weight (kg)", "Layer"]
        hdr_row = detail_hdr_row + 1
        for ci2, h in enumerate(hdrs, 1):
            hdr_cell(ws, hdr_row, ci2, h)

        for bi, b in enumerate(c["placed_boxes"], 1):
            r = hdr_row + bi
            data_cell(ws, r, 1, bi)
            data_cell(ws, r, 2, b["name"], align=left)
            hex_col = b["color"].lstrip("#")
            fill = PatternFill("solid", fgColor=hex_col) if len(hex_col) == 6 else None
            cc = ws.cell(row=r, column=3, value=b["color"])
            if fill: cc.fill = fill
            cc.font = small_font; cc.alignment = center; cc.border = _thin_border()
            data_cell(ws, r, 4, b["rotation"])
            data_cell(ws, r, 5, b["x"])
            data_cell(ws, r, 6, b["y"])
            data_cell(ws, r, 7, b["z"])
            data_cell(ws, r, 8, f"{b['dx']}×{b['dy']}×{b['dz']}")
            data_cell(ws, r, 9, b["weight"])
            layer = int(b["z"] // 300) + 1  # rough layer estimate
            data_cell(ws, r, 10, f"L{layer}")

        for ci2, w2 in enumerate([5, 28, 12, 12, 10, 10, 10, 22, 12, 8], 1):
            ws.column_dimensions[get_column_letter(ci2)].width = w2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_pdf(data: dict) -> io.BytesIO:
    result = data.get("result", data)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)

    styles = getSampleStyleSheet()
    navy  = colors.HexColor("#1B3A5C")
    blue  = colors.HexColor("#2E6DA4")
    light = colors.HexColor("#E8EEF4")
    dblue = colors.HexColor("#D0E4F7")

    title_style = ParagraphStyle("title", fontSize=20, textColor=navy,
                                 fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=4)
    sub_style   = ParagraphStyle("sub",   fontSize=9,  textColor=colors.grey,
                                 fontName="Helvetica",    alignment=TA_CENTER, spaceAfter=12)
    sec_style   = ParagraphStyle("sec",   fontSize=13, textColor=navy,
                                 fontName="Helvetica-Bold", spaceBefore=14, spaceAfter=6)
    sec2_style  = ParagraphStyle("sec2",  fontSize=10, textColor=navy,
                                 fontName="Helvetica-Bold", spaceBefore=8, spaceAfter=4)

    story = []
    story.append(Paragraph("LOAD OPTIMISATION REPORT", title_style))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M')}", sub_style))
    story.append(HRFlowable(width="100%", thickness=1.5, color=navy))
    story.append(Spacer(1, 8))

    # KPI row
    kpi_data = [
        ["Total Boxes", "Total Weight", "Total Volume", "Containers Used", "Unplaced"],
        [str(result["total_boxes"]),
         f"{result['total_weight']:,.1f} kg",
         f"{result['total_volume']/1e9:.4f} m\u00b3",
         str(result["num_containers"]),
         str(len(result.get("unplaced", [])))],
    ]
    kpi_table = Table(kpi_data, colWidths=[54*mm]*5)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), navy),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,0), 9),
        ("BACKGROUND", (0,1), (-1,1), light),
        ("FONTNAME",   (0,1), (-1,1), "Helvetica-Bold"),
        ("FONTSIZE",   (0,1), (-1,1), 11),
        ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("GRID",       (0,0), (-1,-1), 0.5, colors.white),
        ("TOPPADDING",  (0,0), (-1,-1), 7),
        ("BOTTOMPADDING",(0,0), (-1,-1), 7),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 12))

    # ── Container Breakdown ──────────────────────────────────────
    story.append(Paragraph("Container Breakdown", sec_style))
    c_hdrs = ["#", "Container", "Dimensions (L\u00d7W\u00d7H mm)", "Max Wt (kg)",
              "Loaded Wt (kg)", "Wt Util %", "Vol Util %", "Boxes", "Status"]
    c_rows = [c_hdrs]
    for i, c in enumerate(result["containers"], 1):
        cd = c["container_dims"]
        wt_u = c["utilization_wt"]
        status = "Heavy Load" if wt_u > 90 else ("Moderate" if wt_u > 60 else "Light")
        c_rows.append([
            str(i), c["container_name"],
            f"{cd['l']}\u00d7{cd['w']}\u00d7{cd['h']}",
            f"{c['max_wt']:,}", f"{c['total_weight']:,.1f}",
            f"{c['utilization_wt']:.1f}%", f"{c['utilization_vol']:.1f}%",
            str(c["box_count"]), status,
        ])
    cw = [8*mm, 38*mm, 44*mm, 22*mm, 24*mm, 20*mm, 20*mm, 16*mm, 22*mm]
    ct = Table(c_rows, colWidths=cw, repeatRows=1)
    ct.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), navy),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, light]),
        ("GRID",       (0,0), (-1,-1), 0.4, colors.HexColor("#CCCCCC")),
        ("TOPPADDING",  (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0), (-1,-1), 5),
    ]))
    story.append(ct)
    story.append(Spacer(1, 14))

    # ── What's in Each Container ──────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=1, color=blue))
    story.append(Paragraph("What's in Each Container", sec_style))

    for cidx, c in enumerate(result["containers"], 1):
        cd = c["container_dims"]
        container_block = []

        # Container header
        hdr_text = (f"Container {cidx}: {c['container_name']}   |   "
                    f"{c['box_count']} boxes   |   "
                    f"Loaded: {c['total_weight']:,.1f} kg / {c['max_wt']:,} kg   |   "
                    f"Wt: {c['utilization_wt']:.1f}%  Vol: {c['utilization_vol']:.1f}%")
        hdr_para = ParagraphStyle("chdr", fontSize=9, textColor=colors.white,
                                  fontName="Helvetica-Bold", alignment=TA_LEFT,
                                  leftIndent=4, spaceBefore=0, spaceAfter=0)
        hdr_row_data = [[Paragraph(hdr_text, hdr_para)]]
        hdr_tbl = Table(hdr_row_data, colWidths=[214*mm])
        hdr_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), blue),
            ("TOPPADDING",  (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0), (-1,-1), 5),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
        ]))
        container_block.append(hdr_tbl)

        # Box contents table
        contents = _container_contents_summary(c)
        cont_hdrs = ["Box / Item Name", "Qty", "Weight Each (kg)", "Total Weight (kg)", "% of Container Load"]
        cont_rows = [cont_hdrs]
        for box_name, info in contents:
            qty   = info["count"]
            tw    = info["weight"]
            wt_ea = round(tw / qty, 3) if qty else 0
            pct   = round(tw / c["total_weight"] * 100, 1) if c["total_weight"] else 0
            cont_rows.append([box_name, str(qty), f"{wt_ea:.3f}", f"{tw:,.2f}", f"{pct:.1f}%"])

        cont_cw = [74*mm, 20*mm, 36*mm, 42*mm, 42*mm]
        cont_t = Table(cont_rows, colWidths=cont_cw, repeatRows=1)
        cont_t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), dblue),
            ("TEXTCOLOR",  (0,0), (-1,0), navy),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,-1), 8),
            ("ALIGN",      (0,0), (0,-1), "LEFT"),
            ("ALIGN",      (1,0), (-1,-1), "CENTER"),
            ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F4F8FC")]),
            ("GRID",       (0,0), (-1,-1), 0.4, colors.HexColor("#CCCCCC")),
            ("TOPPADDING",  (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",(0,0), (-1,-1), 4),
            ("LEFTPADDING", (0,0), (0,-1), 6),
        ]))
        container_block.append(cont_t)
        container_block.append(Spacer(1, 10))

        story.append(KeepTogether(container_block))

    # ── Per-container box placement details ──────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=1, color=blue))
    story.append(Paragraph("Full Box Placement Details", sec_style))

    for cidx, c in enumerate(result["containers"], 1):
        story.append(Spacer(1, 6))
        story.append(Paragraph(
            f"Container {cidx}: {c['container_name']} — {c['box_count']} boxes, "
            f"{c['total_weight']:,.1f} kg loaded, Wt {c['utilization_wt']:.1f}% / Vol {c['utilization_vol']:.1f}%",
            sec2_style))

        b_hdrs = ["#", "Box Name", "Rotation", "X", "Y", "Z", "dx", "dy", "dz", "Wt (kg)"]
        b_rows = [b_hdrs]
        for bi, b in enumerate(c["placed_boxes"], 1):
            b_rows.append([str(bi), b["name"], b["rotation"],
                           str(b["x"]), str(b["y"]), str(b["z"]),
                           str(b["dx"]), str(b["dy"]), str(b["dz"]),
                           f"{b['weight']:.1f}"])

        bw = [8*mm, 38*mm, 18*mm, 18*mm, 18*mm, 18*mm, 16*mm, 16*mm, 16*mm, 16*mm]
        bt = Table(b_rows, colWidths=bw, repeatRows=1)
        bt.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), blue),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,-1), 7.5),
            ("ALIGN",      (0,0), (-1,-1), "CENTER"),
            ("ALIGN",      (1,0), (1,-1), "LEFT"),
            ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, light]),
            ("GRID",       (0,0), (-1,-1), 0.4, colors.HexColor("#CCCCCC")),
            ("TOPPADDING",  (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",(0,0), (-1,-1), 4),
            ("LEFTPADDING", (1,0), (1,-1), 4),
        ]))
        story.append(bt)

    doc.build(story)
    buf.seek(0)
    return buf